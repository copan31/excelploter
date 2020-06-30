import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import Reference, LineChart, AreaChart, BarChart
import os

class ExcelPloter:
    def __init__(self, csvfile):
        self._df = pd.read_csv(csvfile)
        self._df["DateTime"] = pd.to_datetime(self._df["公表_年月日"])
        self._df = self._df.set_index("DateTime")
        self._file = "test.xlsx"
        print(self._df)

    def _plot(self, sheet, chart, title, xlabel, ylabel, ymax = 0, start_col = 0, end_col = 0, position = 1):
        if start_col == 0:
            start_col = 2 # 指定がなければデータが始まる2列目
        if end_col == 0:
            end_col = sheet.max_column # 指定がなければ最後の列

        # plot on the sheet
        chart.title = title
        chart.x_axis.title = xlabel
        chart.y_axis.title = ylabel
        chart.y_axis.scaling.min = 0
        if ymax != 0:
            chart.y_axis.scaling.max = ymax # 指定があるならy軸の最大値を設定
        chart.legend.position = 'r' # 凡例は右に設定
        chart.add_data(Reference(sheet, min_col = start_col, max_col = end_col, min_row = 1, max_row = sheet.max_row), titles_from_data = True)
        chart.set_categories(Reference(sheet, min_col = 1, max_col = 1, min_row = 2, max_row = sheet.max_row)) # ラベル
        sheet.add_chart(chart, "A" + str(position))

    def reload(self):
        self._df = pd.read_csv(csvfile)

    def print(self):
        print(self._df)

    def plot_line(self, sheetname, column, value):
        # 必要な値と項目を残した時系列データを作成
        df = self._df[[value, column]].pivot_table(index = "DateTime", values = value, columns = column, fill_value = 0)

        with pd.ExcelWriter(self._file) as writer:
            # excelファイルがあるなら追加書きする
            if os.path.isfile(self._file):
                writer.book = load_workbook(self._file)
            # write the df on excel at once
            df.to_excel(writer, sheet_name = sheetname)

            # plot line chart
            chart = LineChart()            
            self._plot(writer.book[sheetname], chart, title = "線グラフ", xlabel = "日付", ylabel = "回数")

    def plot_area(self, sheetname, column, value):
        # 必要な値と項目を残した時系列データを作成
        df = self._df[[value, column]].pivot_table(index = "DateTime", values = value, columns = column, fill_value = 0)

        with pd.ExcelWriter(self._file) as writer:
            # excelファイルがあるなら追加書きする
            if os.path.isfile(self._file):
                writer.book = load_workbook(self._file)

            # write the df on excel at once
            df.to_excel(writer, sheet_name = sheetname)

            # plot Area chart
            chart = AreaChart()
            chart.grouping = "stacked"
            self._plot(writer.book[sheetname], chart, title = "面グラフ", xlabel = "日付", ylabel = "回数")

    def plot_stack(self, sheetname, column, value, unit):
        df = self._df[[value, column]].pivot_table(index = "DateTime", values = value, columns = column, fill_value = 0)
        df = df.resample(unit).sum()

        # self._df.to_csv("a.csv", index = True, header = True, encoding='utf_8_sig')

        with pd.ExcelWriter(self._file) as writer:
            # excelファイルがあるなら追加書きする
            if os.path.isfile(self._file):
                writer.book = load_workbook(self._file)

            # write the df on excel at once
            df.to_excel(writer, sheet_name = sheetname)

            # plot Area chart
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "stacked"
            chart.overlap = 100
            self._plot(writer.book[sheetname], chart, title = "積み上げ棒グラフ", xlabel = "月", ylabel = "回数")

    def plot_stack_mult(self, sheetname, column, value, unit):
        df = self._df[[value, column]].pivot_table(index = "DateTime", values = value, columns = column, fill_value = 0)
        df = df.resample(unit).sum()

        with pd.ExcelWriter(self._file) as writer:
            # excelファイルがあるなら追加書きする
            if os.path.isfile(self._file):
                writer.book = load_workbook(self._file)

            # write the df on excel at once
            df.to_excel(writer, sheet_name = sheetname)
            sheet = writer.book.active # 対象シートを設定

            # step列数ずつ複数のグラフを作成
            step = 3 # 一つの図に描画する列数
            max_col = sheet.max_column # 最大列数
            chart_num = int(max_col / step) # 図の数
            for i in range(chart_num):
                start = i * step + 2
                end = start + (step - 1)
                pos = i * 18 + 1

                title = "積み上げ棒グラフ" + str(i + 1) # 図ごとにタイトルを変える
                chart = BarChart()
                chart.type = "col"
                chart.grouping = "stacked"
                chart.overlap = 100

                self._plot(sheet, chart, title = title, xlabel = "月", ylabel = "回数", ymax = 100, start_col = start, end_col = end, position = pos)

if __name__ == "__main__":
    csvfile = "130001_tokyo_covid19_patients(1).csv"
    csvfile = "sample_pandas_normal.csv"
    a = ExcelPloter(csvfile)
